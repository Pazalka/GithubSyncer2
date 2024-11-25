import pandas as pd
import os
import xlsxwriter
import openpyxl
from datetime import datetime
import tempfile

def create_output_template(workbook, worksheet):
    """יוצר את תבנית קובץ הפלט"""
    # הגדרת פורמטים
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'bg_color': '#D9D9D9'
    })
    
    green_format = workbook.add_format({
        'bg_color': '#C6E0B4'
    })
    
    blue_format = workbook.add_format({
        'bg_color': '#BDD7EE'
    })
    
    yellow_format = workbook.add_format({
        'bg_color': '#FFE699'
    })

    # Row 1 - חותרות חברות
    worksheet.write(0, 0, 'חברה', header_format)
    worksheet.write(0, 1, 'רומי')
    worksheet.write(0, 2, 'יובל') 
    worksheet.write(0, 3, 'החזקות')
    worksheet.write(0, 4, 'החזקות')
    worksheet.write(0, 5, 'נדלן')
    worksheet.write(0, 6, 'נדלן')
    worksheet.write(0, 7, 'נדלן')
    worksheet.write(0, 8, 'יובל פרטי')
    worksheet.write(0, 9, 'יובל פרטי')
    
    # Row 2 - מסגרות
    worksheet.write(1, 0, 'מסגרות', yellow_format)
    worksheet.write(1, 1, '-')
    worksheet.write(1, 2, '-')
    worksheet.write(1, 3, 20000)
    worksheet.write(1, 4, '-')
    worksheet.write(1, 5, 100000)
    worksheet.write(1, 6, 20000)
    worksheet.write(1, 7, 50000)
    worksheet.write(1, 8, 70500)
    worksheet.write(1, 9, 0)
    # Calculate total of numeric values, replacing '-' with 0
    total = sum(0 if x == '-' else x for x in [0, 0, 20000, 0, 100000, 20000, 50000, 70500, 0])
    worksheet.write(1, 10, total)

    # Row 3 - שם הבנק
    worksheet.write(2, 0, 'שם הבנק', header_format)
    worksheet.write(2, 1, 'מזרחי')
    worksheet.write(2, 2, 'לאומי')
    worksheet.write(2, 3, 'פועלים')
    worksheet.write(2, 4, 'מזרחי')
    worksheet.write(2, 5, 'מזרחי')
    worksheet.write(2, 6, 'דיסקונט')
    worksheet.write(2, 7, 'פועלים')
    worksheet.write(2, 8, 'פועלים')
    worksheet.write(2, 9, 'פועלים')
    
    # Row 4 - מספר ח-ן
    account_numbers = ['193744', '4089044', '31324', '177315', '172615', '153771129', '313222', '409937', '55533']
    worksheet.write(3, 0, 'מספר ח-ן', header_format)
    for col, number in enumerate(account_numbers, start=1):
        worksheet.write(3, col, number, green_format)
    
    worksheet.right_to_left()
    return account_numbers

def find_account_number(sheet):
    """מוצא את מספר החשבון בגיליון"""
    bank_keywords = ['מזרחי', 'פועלים', 'דיסקונט', 'לאומי']
    
    for row in range(1, sheet.max_row + 1):
        cell_value = str(sheet.cell(row=row, column=2).value or '')
        
        # מחפש את כל סוגי הבנקים
        if any(bank in cell_value for bank in bank_keywords):
            # מחפש מספר בתוך הטקסט
            words = cell_value.split()
            for word in words:
                # בודק אם המילה מכילה רק מספרים ולפחות 5 ספרות
                if word.isdigit() and len(word) >= 5:
                    return word
            
    raise ValueError("לא נמצא מספר חשבון בקובץ")

def find_output_column(account_number, account_numbers):
    """מוצא את העמודה המתאימה בקובץ הפלט"""
    for col, acc in enumerate(account_numbers, start=1):
        if acc == account_number:
            return col
    raise ValueError(f"לא נמצא חשבון מתאים {account_number}")

def collect_all_dates(bank_files):
    """אוסף את כל התאריכים מכל הקבצים"""
    all_dates = set()
    
    for input_file in bank_files:
        wb = openpyxl.load_workbook(input_file)
        sheet = wb.active
        
        for row in range(1, sheet.max_row + 1):
            date_cell = sheet.cell(row=row, column=1).value
            if isinstance(date_cell, str):
                try:
                    date = datetime.strptime(date_cell, '%d/%m/%y')
                    all_dates.add(date)
                except ValueError:
                    continue
        
        wb.close()
    
    return sorted(all_dates)

def process_excel_files(file_paths):
    """
    Process the uploaded Excel files and return the path to the output file.
    """
    # Create a temporary file for the output
    temp_output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    
    try:
        workbook = xlsxwriter.Workbook(temp_output.name)
        worksheet = workbook.add_worksheet()
        
        account_numbers = create_output_template(workbook, worksheet)
        
        # Collect all dates from the uploaded files
        all_dates = collect_all_dates(file_paths)
        
        # Set up formats
        date_format = workbook.add_format({'num_format': 'dd/mm/yy'})
        number_format = workbook.add_format({'num_format': '#,##0.00'})
        
        # Write sum column header
        worksheet.write(3, 10, 'סה"כ', workbook.add_format({
            'bold': True,
            'align': 'center',
            'bg_color': '#D9D9D9'
        }))
        
        # Write dates and prepare sum formulas
        for row, date in enumerate(all_dates, start=4):
            worksheet.write_datetime(row, 0, date, date_format)
            sum_formula = f'=SUM(B{row+1}:J{row+1})'
            worksheet.write_formula(row, 10, sum_formula, number_format)
        
        # Process each bank file
        for input_file in file_paths:
            try:
                wb = openpyxl.load_workbook(input_file)
                sheet = wb.active
                
                account_number = find_account_number(sheet)
                output_col = find_output_column(account_number, account_numbers)
                
                # Collect balances
                balances = {}
                for row in range(1, sheet.max_row + 1):
                    date_str = sheet.cell(row=row, column=1).value
                    balance = sheet.cell(row=row, column=10).value
                    if isinstance(date_str, str) and isinstance(balance, (int, float)):
                        try:
                            date = datetime.strptime(date_str, '%d/%m/%y')
                            balances[date] = balance
                        except ValueError:
                            continue
                
                # Update output file with balances
                for row, date in enumerate(all_dates, start=4):
                    if date in balances:
                        worksheet.write_number(row, output_col, balances[date], number_format)
                
                wb.close()
                
            except Exception as e:
                raise ValueError(f"שגיאה בעיבוד הקובץ: {str(e)}")
                
        workbook.close()
        return temp_output.name
        
    except Exception as e:
        if os.path.exists(temp_output.name):
            os.remove(temp_output.name)
        raise e
